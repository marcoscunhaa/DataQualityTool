# -*- coding: utf-8 -*-
"""
Created on Thu Mar 26 17:35:52 2026
@author: marcos cunha
"""

import os
import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Font

#Função que carrega o arquivo .xlsx, .xls, .csv ou .txt
def load_data(path):
    if not os.path.exists(path):
        print(f"❌ Error: File {path} not found.")
        return None

    extension = os.path.splitext(path)[1].lower()

    try:
        if extension in ['.xlsx', '.xls']:
            return pd.read_excel(path)
        elif extension in ['.csv', '.txt']:
            # Tentativa de carregar com diferentes encodings comuns no Brasil/Windows
            for enc in ['utf-8', 'latin1', 'iso-8859-1', 'cp1252']:
                try:
                    return pd.read_csv(path, sep=None, engine='python', encoding=enc)
                except:
                    continue
            return None
        else:
            print("❌ Unsupported file format.")
            return None
    except Exception as e:
        print(f"❌ Critical loading error: {e}")
        return None

#Função que scaneia os erros da planilha
def scan_errors(df, export_path="result/report.xlsx"):
    print("📊 Analyzing Data Quality...")
    report = []
    rows, cols = df.shape
    total_cells = rows * cols
    total_errors = 0

    #Verificação de nulos
    nulls = df.isnull().sum()
    for col, count in nulls.items():
        if count > 0:
            report.append({"type": "missing", "column": col, "count": int(count)})
            total_errors += count

    #Coluna com tipagem mista
    for col in df.columns:
        types = df[col].dropna().apply(type).unique()
        if len(types) > 1:
            report.append({
                "type": "mixed data types in column",
                "column": col,
                "count": len(types)
            })
            total_errors += 1

    #Linhas completas duplicadas
    dups = df.duplicated().sum()
    if dups > 0:
        report.append({"type": "duplicate rows", "column": "ALL", "count": int(dups)})
        total_errors += dups

    #Identificador único duplicados
    id_cols = [c for c in df.columns if 'id' in str(c).lower()]
    for col in id_cols:
        dup_count = df[col].dropna().duplicated().sum()
        if dup_count > 0:
            report.append({"type": "duplicate id", "column": col, "count": int(dup_count)})
            total_errors += dup_count

    #Valores numéricos formatado em texto
    terms_price = ['price', 'value', 'cost', 'amount', 'total', 'sale', 'unit']
    price_cols = [c for c in df.columns if any(t in str(c).lower() for t in terms_price)]
    
    for col in price_cols:
        if df[col].dtype == 'object':
            report.append({
                "type": "price column is text or mixed type",
                "column": col,
                "count": int(df[col].notna().sum())
            })
            total_errors += 1

        temp_raw = df[col].astype(str).str.strip()
        invalid_mask = temp_raw.str.contains(r'[^0-9.\-]', na=False) & (temp_raw != 'nan') & (temp_raw != '')
        invalid_count = invalid_mask.sum()

        if invalid_count > 0:
            report.append({
                "type": "invalid characters in price field",
                "column": col,
                "count": int(invalid_count)
            })
            total_errors += invalid_count

        temp_numeric = pd.to_numeric(temp_raw.str.replace(r'[^0-9.\-]', '', regex=True), errors='coerce')
        if temp_numeric.notna().sum() > 0:
            q1 = temp_numeric.quantile(0.25)
            q3 = temp_numeric.quantile(0.75)
            iqr = q3 - q1
            outliers = ((temp_numeric < (q1 - 1.5 * iqr)) | (temp_numeric > (q3 + 1.5 * iqr))).sum()
            if outliers > 0:
                report.append({"type": "outliers detected (IQR)", "column": col, "count": int(outliers)})
                total_errors += outliers

    #Datas fora do formado US MM/DD/YYYY
    terms_date = ['date', 'timestamp', 'created', 'day', 'month', 'year']
    date_cols = [c for c in df.columns if any(t in str(c).lower() for t in terms_date)]
    for col in date_cols:
        invalid_count = (pd.to_datetime(df[col], errors='coerce', dayfirst=True).isna()& df[col].notna()).sum()
        if invalid_count > 0:
            report.append({"type": "the date format is not US.", "column": col, "count": int(invalid_count)})
            total_errors += invalid_count

    #Inconsistência de texto
    for col in df.columns:
        if pd.api.types.is_string_dtype(df[col]) or df[col].dtype == 'object':
            series = df[col].dropna().astype(str)
            if len(series) == 0 or series.nunique() / len(series) > 0.9:
                continue

            temp_df = pd.DataFrame({"original": series, "normalized": series.str.strip().str.lower()})
            grouped = temp_df.groupby("normalized")
            inconsistent_count = sum(len(group) for _, group in grouped if len(group["original"].unique()) > 1)

            if inconsistent_count > 0:
                report.append({"type": "text inconsistency (case/space)", "column": col, "count": inconsistent_count})
                total_errors += inconsistent_count

    #Contatos não formatado em texto
    contact_terms = ['contact', 'phone', 'email', 'telefone', 'celular', 'mobile']
    contact_cols = [c for c in df.columns if any(term in str(c).lower() for term in contact_terms)]
    for col in contact_cols:
        if not (pd.api.types.is_string_dtype(df[col]) or df[col].dtype == 'object'):
            val_count = int(df[col].notna().sum())
            report.append({"type": "contact column not text", "column": col, "count": val_count})
            total_errors += val_count

    #Valores negativos em colunas criticas
    terms_quantity = ['quantity', 'qty', 'amount', 'count', 'units']
    quantity_cols = [c for c in df.columns if any(term in str(c).lower() for term in terms_quantity)]
    for col in quantity_cols:
        if pd.api.types.is_numeric_dtype(df[col]):
            negatives = (df[col] < 0).sum()
            if negatives > 0:
                report.append({"type": "negative values", "column": col, "count": int(negatives)})
                total_errors += negatives

    #Verificação de abreviações em grupos de dados ou categorias
    terms_category = ['category', 'categor', 'cat', 'type', 'class']
    category_cols = [c for c in df.columns if any(term in str(c).lower() for term in terms_category)]

    for col in category_cols:
        series = df[col].dropna().astype(str).str.strip()
        if series.empty:
            continue
        counts = series.value_counts()

        threshold = max(2, len(series) * 0.05)
        rare_mask = (counts < threshold) | (counts.index.str.endswith('.'))
        rare_values = counts[rare_mask]

        if not rare_values.empty:
            report.append({
                "type": "potential abbreviations",
                "column": col,
                "count": int(rare_values.sum())
            })
            total_errors += rare_values.sum()

    # Cálculo do score de qualidade da planilhha original
    quality_score = max(0, 100 - ((total_errors / total_cells) * 100)) if total_cells > 0 else 100

    if quality_score >= 90:
        quality_label = "Excellent"
    elif quality_score >= 75:
        quality_label = "Good"
    elif quality_score >= 50:
        quality_label = "Regular"
    else:
        quality_label = "Poor"

    summary_df = pd.DataFrame({
        "Metric": ["Total Rows", "Total Columns", "Total Cells", "Total Errors Found", "Error Rate (%)", "Data Quality Score (%)", "Quality Level"],
        "Value": [
            rows, cols, total_cells, total_errors,
            f"{(total_errors/total_cells)*100:.2f}%" if total_cells > 0 else "0%",
            f"{quality_score:.2f}%",
            quality_label
        ]
    })

    report_df = pd.DataFrame(report)
    if not report_df.empty:
        report_df["percentage"] = (report_df["count"] / rows * 100).apply(lambda x: f"{x:.2f}%")
        report_df = report_df[['column', 'type', 'count', 'percentage']]

    #Exportação do relatório de auditoria
    try:
        output_dir = os.path.dirname(export_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            # Padronização do texto
            summary_df.columns = summary_df.columns.str.title()
            if not report_df.empty:
                report_df.columns = report_df.columns.str.title()
                
            summary_df.to_excel(writer, sheet_name="Quality Summary", index=False)
            if not report_df.empty:
                report_df.to_excel(writer, sheet_name="Error Details", index=False)
            
            summary_df.to_excel(writer, sheet_name="Quality Summary", index=False)
            if not report_df.empty:
                report_df.to_excel(writer, sheet_name="Error Details", index=False)

            # Estilo do texto
            ws_sum = writer.sheets['Quality Summary']
            for row in ws_sum.iter_rows(min_row=2, max_row=8, min_col=2, max_col=2):
                for cell in row:
                    if "Poor" in str(cell.value):
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif "Excellent" in str(cell.value):
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                header_row = ws[1]
                for cell in header_row:
                    cell.font = Font(bold=True)

            # Alinhamento geral
            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left')

        print(f"📈 Quality Score: {quality_score:.2f}% ({quality_label})")
        print(f"📁 Report saved to: {export_path}")

    except Exception as e:
        print(f"❌ Error saving Excel: {e}")

    return report_df

# Execução do script
file_path = "Walmart Inventory.csv"
df = load_data(file_path)
if df is not None:
    report = scan_errors(df)