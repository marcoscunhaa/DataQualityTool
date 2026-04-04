# -*- coding: utf-8 -*-
"""
Created on Thu Mar 26 17:35:52 2026

@author: marcos cunha
"""

import os
import pandas as pd
import numpy as np
import re
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook

# Função de carregar o arquivo
def load_data(path):
    if not os.path.exists(path):
        print(f"❌ Error: File {path} not found.")
        return None

    extension = os.path.splitext(path)[1].lower()

    try:
        if extension in ['.xlsx', '.xls']:
            return pd.read_excel(path)
        elif extension in ['.csv', '.txt']:
            for enc in ['utf-8', 'latin1', 'iso-8859-1', 'cp1252']:
                try:
                    return pd.read_csv(path, sep=None, engine='python', encoding=enc)
                except:
                    continue
            return None
        else:
            return None
    except Exception as e:
        print(f"❌ Critical loading error: {e}")
        return None


# Detecta formato de data
def detect_date_format(series):
    sample = series.dropna().astype(str).head(20)

    day_first_count = 0
    month_first_count = 0

    for val in sample:
        try:
            dt_auto = pd.to_datetime(val, errors='coerce', format='mixed')

            if pd.isna(dt_auto):
                continue

            if '/' in val or '-' in val:
                parts = val.replace('-', '/').split('/')

                if len(parts) == 3:
                    a, b, c = parts

                    if int(a) > 12:
                        day_first_count += 1
                    elif int(b) > 12:
                        month_first_count += 1

        except:
            continue

    return day_first_count >= month_first_count


# Função principal
def clean_data(df, output_path="result/correction.xlsx"):
    df_original = df.copy()
    df = df.copy()

    print("\n🧹 STARTING DATA CLEANING...")

    initial_rows = len(df)

    #Configuração de variaveis
    date_cols = [c for c in df.columns if 'date' in c.lower()]
    currency_cols = [c for c in df.columns if any(k in c.lower() for k in ['price', 'value', 'amount', 'total', 'sales', 'cost'])]
    id_cols = [c for c in df.columns if 'id' in c.lower()]
    discount_cols = [c for c in df.columns if 'discount' in c.lower()]
    critical_keywords = ['date', 'price', 'product', 'id', 'quantity']
    critical_cols = [c for c in df.columns if any(k in c.lower() for k in critical_keywords)]
    
    #Configuração para ajuste de case(Maiusculas ou Minusculas)
    text_standardization_terms = ['country']
    text_replacements = {
        'Usa': 'USA',
    }
    
    #Mapeamento para grupos ou categorias.
    CATEGORY_MAPPING = {
        'Appl.': 'Appliances',
        'App': 'Apparel',
        'Apparel': 'Apparel',
        'Cloth.': 'Clothing',
        'Clothing': 'Clothing',
        'Furn.': 'Furniture',
        'Furn': 'Furniture',
        'Furniture': 'Furniture',
        'Electr.': 'Electronics',
        'Electronics': 'Electronics',
        'Electronic Devices': 'Electronics',
        'Home': 'Home'
    }
    
    #Converte colunas de contato em string
    contact_keywords = ['contact', 'phone', 'email', 'tel']
    contact_cols = [c for c in df.columns if any(k in c.lower() for k in contact_keywords)]

    for col in contact_cols:
        converted_count = 0

        def convert_value(x):
            nonlocal converted_count

            if isinstance(x, float) and not pd.isna(x):
                converted_count += 1
                return str(int(x))
            elif pd.isna(x):
                return np.nan
            return str(x)

        df[col] = df[col].apply(convert_value)

        df[col] = df[col].replace(['nan', 'None', 'NaN'], np.nan)
        df[col] = df[col].fillna("").astype(str).str.strip()
        
    keep_mask = pd.Series(True, index=df.index)

    #Remove linhas duplicadas
    if id_cols:
        is_duplicate = df.duplicated(subset=id_cols, keep='first')
    else:
        is_duplicate = df.duplicated(keep='first')

    keep_mask &= ~is_duplicate

    #Tratamento de moedas
    for col in currency_cols:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].str.replace(r'[^\d.,]', '', regex=True)

        def parse_currency(value):
            try:
                if pd.isna(value):
                    return np.nan

                value = str(value).strip()

                if value == '' or value.lower() == 'nan':
                    return np.nan

                value = re.sub(r'[^\d,.\-]', '', value)

                if ',' in value and '.' in value:
                    if value.rfind(',') > value.rfind('.'):
                        value = value.replace('.', '').replace(',', '.')
                    else:
                        value = value.replace(',', '')
                elif ',' in value:
                    value = value.replace(',', '.')

                return float(value)

            except:
                return np.nan

        df[col] = df[col].apply(parse_currency)


    #Converte o formado da data para US
    for col in date_cols:
        df[col] = df[col].astype(str).str.strip()

        dayfirst_detected = detect_date_format(df[col])

        df[col] = pd.to_datetime(
            df[col],
            format='mixed',
            dayfirst=dayfirst_detected,
            errors="coerce"
        )

        df[col] = df[col].dt.strftime('%m/%d/%Y')

    #Trata de descontos financeiros que estão em nulo
    for col in discount_cols:
        df[col] = df[col].fillna(0)

    #Converte texto nulo em Unknown
    text_cols = df.select_dtypes(include=['object', 'string']).columns
    non_critical_text = [c for c in text_cols if c not in critical_cols]

    for col in non_critical_text:
        df[col] = df[col].replace(['nan', 'NaN', 'None', 'null', ""], np.nan)
        df[col] = df[col].fillna("Unknown")

    #Tratatemento de colunas com incompatibilidade de case(Maiuscula ou Minuscula)
    target_text_cols = [c for c in df.columns if any(term in c.lower() for term in text_standardization_terms)]

    for col in target_text_cols:
        df[col] = (
            df[col]
            .astype(str)
            .str.strip()
            .str.lower()
            .str.title()
            .replace(text_replacements)
        )

    #Converte as categorias ou grupos abreviados
    category_cols = [c for c in df.columns if any(term in c.lower() for term in ['category','categor','cat','type','class'])]

    for col in category_cols:

        def normalize_category(x):
            if pd.isna(x):
                return x
            x_str = str(x).strip()
            return CATEGORY_MAPPING.get(x_str, x_str)

        df[col] = df[col].apply(normalize_category)

    # FINAL
    df_final = df[keep_mask].copy()
    df_discarded = df_original[~keep_mask].copy()

    print(f"📉 Rows before: {initial_rows}")
    print(f"✅ Rows after cleaning: {len(df_final)}")
    print(f"❌ Rows discarded: {len(df_discarded)}")

    # Export
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='Cleaned Data', index=False)
        df_discarded.to_excel(writer, sheet_name='Discarded Data', index=False)

    #Formatação do excel
    wb = load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Formatar cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left')

            # Title Case no header
            if isinstance(cell.value, str):
                cell.value = cell.value.strip().title()

    wb.save(output_path)

    print(f"📁 File saved as: {output_path}")
    return df_final


# Execução
file_path = "Walmart Inventory.csv"
df = load_data(file_path)

if df is not None:
    df_clean = clean_data(df)